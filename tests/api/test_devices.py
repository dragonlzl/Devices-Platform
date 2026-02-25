import os
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.db import init_db
from backend.main import app


@pytest.fixture()
def client():
    os.environ["APP_DB_FILE"] = "data/apitest.db"
    db_path = Path(__file__).resolve().parents[2] / "data" / "apitest.db"
    if db_path.exists():
        db_path.unlink()
    init_db()
    with TestClient(app) as client:
        yield client


def create_vendor_system_version(client):
    vendor = client.post("/api/vendors", json={"name": "Apple"})
    assert vendor.status_code == 200
    system = client.post("/api/systems", json={"name": "iOS"})
    assert system.status_code == 200
    systems = client.get("/api/systems?include_versions=1").json()["items"]
    system_id = systems[0]["id"]
    version = client.post(f"/api/systems/{system_id}/versions", json={"version": "17.0"})
    assert version.status_code == 200
    vendors = client.get("/api/vendors").json()["items"]
    vendor_id = vendors[0]["id"]
    systems = client.get("/api/systems?include_versions=1").json()["items"]
    version_id = systems[0]["versions"][0]["id"]
    return vendor_id, system_id, version_id


def test_device_borrow_return_flow(client):
    vendor_id, system_id, version_id = create_vendor_system_version(client)
    payload = {
        "model": "iPhone 14",
        "status": "正常",
        "type": "手机",
        "vendor_id": vendor_id,
        "system_id": system_id,
        "system_version_id": version_id,
        "resolution": "1170x2532",
        "arch": "arm64",
        "cpu": "A15",
        "boot_password": "1234",
        "notes": "测试机",
    }
    resp = client.post("/api/devices", json=payload)
    assert resp.status_code == 200

    devices = client.get("/api/devices").json()["items"]
    device_id = devices[0]["id"]

    borrow = client.post(
        f"/api/devices/{device_id}/borrow",
        json={
            "borrower_name": "Alice",
            "expected_return_at": "2099-12-31T10:00:00+00:00",
        },
    )
    assert borrow.status_code == 200

    borrowed = client.get("/api/devices").json()["items"][0]
    assert borrowed["loan_status"] == "pending"
    assert borrowed["borrower_name"] == "Alice"
    assert borrowed["borrowed_at"] is None

    requests = client.get("/api/borrow-requests").json()["items"]
    request_id = requests[0]["id"]
    approve = client.post(f"/api/borrow-requests/{request_id}/approve")
    assert approve.status_code == 200

    borrowed = client.get("/api/devices").json()["items"][0]
    assert borrowed["loan_status"] == "borrowed"
    assert borrowed["borrowed_at"] is not None

    extend = client.post(
        f"/api/devices/{device_id}/extend",
        json={"expected_return_at": "2099-12-31T12:00:00+00:00"},
    )
    assert extend.status_code == 200

    returned = client.post(f"/api/devices/{device_id}/return")
    assert returned.status_code == 200

    final = client.get("/api/devices").json()["items"][0]
    assert final["loan_status"] == "available"
    assert final["borrower_name"] is None

    records = client.get("/api/borrow-records").json()["items"]
    assert records[0]["status"] == "returned"
    assert records[0]["returned_at"] is not None


def test_borrow_request_cancel(client):
    vendor_id, system_id, version_id = create_vendor_system_version(client)
    client.post(
        "/api/devices",
        json={
            "model": "Pixel 8",
            "status": "正常",
            "type": "手机",
            "vendor_id": vendor_id,
            "system_id": system_id,
            "system_version_id": version_id,
        },
    )
    device_id = client.get("/api/devices").json()["items"][0]["id"]
    borrow = client.post(
        f"/api/devices/{device_id}/borrow",
        json={
            "borrower_name": "Bob",
            "expected_return_at": "2099-12-31T10:00:00+00:00",
        },
    )
    assert borrow.status_code == 200
    pending = client.get("/api/borrow-requests?status=pending")
    assert pending.status_code == 200
    assert len(pending.json()["items"]) == 1
    request_id = client.get("/api/borrow-requests").json()["items"][0]["id"]
    cancel = client.post(f"/api/borrow-requests/{request_id}/cancel")
    assert cancel.status_code == 200
    pending = client.get("/api/borrow-requests?status=pending").json()["items"]
    assert pending == []
    device = client.get("/api/devices").json()["items"][0]
    assert device["loan_status"] == "available"


def test_borrow_rejected_for_unregistered_status(client):
    vendor_id, system_id, version_id = create_vendor_system_version(client)
    client.post(
        "/api/devices",
        json={
            "model": "UnregisteredPhone",
            "status": "未登记借用",
            "type": "手机",
            "vendor_id": vendor_id,
            "system_id": system_id,
            "system_version_id": version_id,
        },
    )
    device_id = client.get("/api/devices").json()["items"][0]["id"]
    borrow = client.post(
        f"/api/devices/{device_id}/borrow",
        json={
            "borrower_name": "Alice",
            "expected_return_at": "2099-12-31T10:00:00+00:00",
        },
    )
    assert borrow.status_code == 400
    assert "未找到该设备的借用人" in borrow.json()["detail"]


def test_vendor_rebind_on_delete(client):
    vendor_id, system_id, version_id = create_vendor_system_version(client)
    client.post(
        "/api/devices",
        json={
            "model": "iPad",
            "status": "正常",
            "type": "平板",
            "vendor_id": vendor_id,
            "system_id": system_id,
            "system_version_id": version_id,
        },
    )
    client.post("/api/vendors", json={"name": "Samsung"})
    vendors = client.get("/api/vendors").json()["items"]
    target_vendor = next(item for item in vendors if item["name"] == "Samsung")

    resp = client.request("DELETE", f"/api/vendors/{vendor_id}", json={"rebind_vendor_id": None})
    assert resp.status_code == 409

    resp = client.request(
        "DELETE",
        f"/api/vendors/{vendor_id}",
        json={"rebind_vendor_id": target_vendor["id"]},
    )
    assert resp.status_code == 200

    devices = client.get("/api/devices").json()["items"]
    assert devices[0]["vendor_name"] == "Samsung"


def test_version_rebind_on_delete(client):
    vendor_id, system_id, version_id = create_vendor_system_version(client)
    client.post(
        "/api/devices",
        json={
            "model": "Pixel",
            "status": "正常",
            "type": "手机",
            "vendor_id": vendor_id,
            "system_id": system_id,
            "system_version_id": version_id,
        },
    )
    client.post(f"/api/systems/{system_id}/versions", json={"version": "18.0"})
    systems = client.get("/api/systems?include_versions=1").json()["items"]
    versions = systems[0]["versions"]
    other_version = next(item for item in versions if item["version"] == "18.0")

    resp = client.request(
        "DELETE", f"/api/versions/{version_id}", json={"rebind_version_id": None}
    )
    assert resp.status_code == 409

    resp = client.request(
        "DELETE",
        f"/api/versions/{version_id}",
        json={"rebind_version_id": other_version["id"]},
    )
    assert resp.status_code == 200

    devices = client.get("/api/devices").json()["items"]
    assert devices[0]["system_version"] == "18.0"


def test_device_query_by_id(client):
    vendor_id, system_id, version_id = create_vendor_system_version(client)
    client.post(
        "/api/devices",
        json={
            "model": "QueryPhone",
            "status": "正常",
            "type": "手机",
            "vendor_id": vendor_id,
            "system_id": system_id,
            "system_version_id": version_id,
        },
    )
    device_id = client.get("/api/devices").json()["items"][0]["id"]
    resp = client.get(f"/api/devices?query={device_id}")
    assert resp.status_code == 200
    items = resp.json()["items"]
    assert len(items) == 1
    assert items[0]["id"] == device_id


def test_change_borrower_updates_request(client):
    vendor_id, system_id, version_id = create_vendor_system_version(client)
    client.post(
        "/api/devices",
        json={
            "model": "ChangeBorrowerPhone",
            "status": "正常",
            "type": "手机",
            "vendor_id": vendor_id,
            "system_id": system_id,
            "system_version_id": version_id,
        },
    )
    device = client.get("/api/devices").json()["items"][0]
    device_id = device["id"]
    borrow = client.post(
        f"/api/devices/{device_id}/borrow",
        json={
            "borrower_name": "Alice",
            "expected_return_at": "2099-12-31T10:00:00+00:00",
        },
    )
    assert borrow.status_code == 200
    request_id = client.get("/api/borrow-requests").json()["items"][0]["id"]
    approve = client.post(f"/api/borrow-requests/{request_id}/approve")
    assert approve.status_code == 200

    change = client.post(
        f"/api/devices/{device_id}/change-borrower",
        json={
            "borrower_name": "Bob",
            "expected_return_at": "2099-12-31T12:00:00+00:00",
        },
    )
    assert change.status_code == 200
    device = client.get("/api/devices").json()["items"][0]
    assert device["borrower_name"] == "Alice"
    change_request = next(
        item for item in client.get("/api/borrow-requests").json()["items"] if item["request_type"] == "change"
    )
    approve_change = client.post(f"/api/borrow-requests/{change_request['id']}/approve")
    assert approve_change.status_code == 200

    updated = client.get("/api/devices").json()["items"][0]
    assert updated["borrower_name"] == "Bob"
    assert updated["expected_return_at"] == "2099-12-31T12:00:00+00:00"

    records = client.get("/api/borrow-records").json()["items"]
    assert records
    changes = records[0].get("borrower_changes", [])
    assert len(changes) == 1
    assert changes[0]["borrower_before"] == "Alice"
    assert changes[0]["borrower_after"] == "Bob"


def test_change_borrower_cancel_does_not_reset_device(client):
    vendor_id, system_id, version_id = create_vendor_system_version(client)
    client.post(
        "/api/devices",
        json={
            "model": "ChangeBorrowerCancelPhone",
            "status": "正常",
            "type": "手机",
            "vendor_id": vendor_id,
            "system_id": system_id,
            "system_version_id": version_id,
        },
    )
    device = client.get("/api/devices").json()["items"][0]
    device_id = device["id"]
    borrow = client.post(
        f"/api/devices/{device_id}/borrow",
        json={
            "borrower_name": "Alice",
            "expected_return_at": "2099-12-31T10:00:00+00:00",
        },
    )
    assert borrow.status_code == 200
    request_id = client.get("/api/borrow-requests").json()["items"][0]["id"]
    approve = client.post(f"/api/borrow-requests/{request_id}/approve")
    assert approve.status_code == 200

    change = client.post(
        f"/api/devices/{device_id}/change-borrower",
        json={
            "borrower_name": "Bob",
            "expected_return_at": "2099-12-31T12:00:00+00:00",
        },
    )
    assert change.status_code == 200
    change_request = next(
        item for item in client.get("/api/borrow-requests").json()["items"] if item["request_type"] == "change"
    )
    cancel = client.post(f"/api/borrow-requests/{change_request['id']}/cancel")
    assert cancel.status_code == 200

    pending = client.get("/api/borrow-requests?status=pending").json()["items"]
    assert pending == []
    updated = client.get("/api/devices").json()["items"][0]
    assert updated["borrower_name"] == "Alice"
    assert updated["expected_return_at"] == "2099-12-31T10:00:00+00:00"
    cancelled = next(
        item for item in client.get("/api/borrow-requests").json()["items"] if item["request_type"] == "change"
    )
    assert cancelled["request_status"] == "cancelled"


def test_change_borrower_rejected_if_pending_request_exists(client):
    vendor_id, system_id, version_id = create_vendor_system_version(client)
    client.post(
        "/api/devices",
        json={
            "model": "ChangeBorrowerBlockPhone",
            "status": "正常",
            "type": "手机",
            "vendor_id": vendor_id,
            "system_id": system_id,
            "system_version_id": version_id,
        },
    )
    device = client.get("/api/devices").json()["items"][0]
    device_id = device["id"]
    borrow = client.post(
        f"/api/devices/{device_id}/borrow",
        json={
            "borrower_name": "Alice",
            "expected_return_at": "2099-12-31T10:00:00+00:00",
        },
    )
    assert borrow.status_code == 200
    request_id = client.get("/api/borrow-requests").json()["items"][0]["id"]
    approve = client.post(f"/api/borrow-requests/{request_id}/approve")
    assert approve.status_code == 200

    first_change = client.post(
        f"/api/devices/{device_id}/change-borrower",
        json={
            "borrower_name": "Bob",
            "expected_return_at": "2099-12-31T12:00:00+00:00",
        },
    )
    assert first_change.status_code == 200

    second_change = client.post(
        f"/api/devices/{device_id}/change-borrower",
        json={
            "borrower_name": "Cindy",
            "expected_return_at": "2099-12-31T13:00:00+00:00",
        },
    )
    assert second_change.status_code == 400
    assert "已有借用人更换申请" in second_change.json()["detail"]
    assert "等待管理员处理" in second_change.json()["detail"]

    pending = client.get("/api/borrow-requests?status=pending&request_type=change").json()["items"]
    assert len(pending) == 1


def test_export_devices(client):
    vendor_id, system_id, version_id = create_vendor_system_version(client)
    client.post(
        "/api/devices",
        json={
            "model": "ExportBorrowPhone",
            "status": "正常",
            "type": "手机",
            "vendor_id": vendor_id,
            "system_id": system_id,
            "system_version_id": version_id,
        },
    )
    device_id = client.get("/api/devices").json()["items"][0]["id"]
    borrow = client.post(
        f"/api/devices/{device_id}/borrow",
        json={
            "borrower_name": "ExportUser",
            "expected_return_at": "2099-12-31T10:00:00+00:00",
        },
    )
    assert borrow.status_code == 200
    request_id = client.get("/api/borrow-requests").json()["items"][0]["id"]
    approve = client.post(f"/api/borrow-requests/{request_id}/approve")
    assert approve.status_code == 200

    resp = client.get("/api/devices/export")
    assert resp.status_code == 200
    wb = load_workbook(filename=BytesIO(resp.content))
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    assert "借用人" in header
    assert "设备型号" in header
    assert "借用状态" in header
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert any(row[1] == "ExportBorrowPhone" for row in rows)
